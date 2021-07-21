import openpyxl 
from src.PersistenceManager import PersistenceManager
from time import sleep
from tkinter import filedialog
from os import mkdir, path, getcwd

pm = PersistenceManager()


class WorkbooktUtil():
    
    def __init__(self):
        self.saveDir = pm.loadSaveDirWorkbooks()
    
    def createWorkbook(self, dataToWrite, workbookName, columnNames):
        try:
            workbook = openpyxl.Workbook() 
            sheet = workbook.create_sheet("Sheet")
            sheet = self.writeData(dataToWrite, sheet, columnNames)
            workbook.save(path.join(self.saveDir, workbookName+".xls"))

        except Exception as erro:
            raise erro

    def writeData(self, dataToWrite, sheetToWrite, columnNames):
        
        try:
            for column, name in enumerate(columnNames):
                sheetToWrite.cell(row=1, column=column + 1, value=name,)

            for row, data in enumerate(dataToWrite):
                for column in range(1, len(data)):
                    sheetToWrite.cell(row=row + 1, column=column, value=data[column])
            
            return sheetToWrite
        
        except Exception as erro:
            raise erro



